import os
import csv
import json
import base64
import uuid
from typing import Any, Dict, List
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font
from fpdf import FPDF
from pathlib import Path

# Configurable storage directory
EXPORT_DIR = Path("exports")
EXPORT_DIR.mkdir(exist_ok=True)

class ExportService:
    @staticmethod
    async def generate_csv(file_id: uuid.UUID, data: List[Dict[str, Any]]) -> str:
        file_path = EXPORT_DIR / f"{file_id}.csv"
        if not data:
            with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                f.write('')
            return str(file_path)

        headers = list(data[0].keys())
        with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=headers, quoting=csv.QUOTE_MINIMAL)
            writer.writeheader()
            for row in data:
                writer.writerow(row)
                
        return str(file_path)

    @staticmethod
    async def generate_excel(file_id: uuid.UUID, data: List[Dict[str, Any]], insights: Dict[str, Any], metadata: Dict[str, Any]) -> str:
        file_path = EXPORT_DIR / f"{file_id}.xlsx"
        wb = openpyxl.Workbook()
        
        # Sheet 1: Data
        ws_data = wb.active
        ws_data.title = "Data"
        if data:
            headers = list(data[0].keys())
            ws_data.append(headers)
            # Bold headers
            for col_num in range(1, len(headers) + 1):
                cell = ws_data.cell(row=1, column=col_num)
                cell.font = Font(bold=True)
            
            for row in data:
                ws_data.append([row.get(h) for h in headers])
            
            # Freeze top row & Add filters
            ws_data.freeze_panes = 'A2'
            ws_data.auto_filter.ref = ws_data.dimensions
            
            # Auto column width
            for col in ws_data.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws_data.column_dimensions[column].width = min(adjusted_width, 50)

        # Sheet 2: Insights
        ws_insights = wb.create_sheet(title="Insights")
        ws_insights.append(["Section", "Content"])
        ws_insights.cell(row=1, column=1).font = Font(bold=True)
        ws_insights.cell(row=1, column=2).font = Font(bold=True)
        ws_insights.column_dimensions['A'].width = 20
        ws_insights.column_dimensions['B'].width = 100
        
        if insights:
            ws_insights.append(["Summary", insights.get("summary", "")])
            for i, finding in enumerate(insights.get("key_findings", [])):
                ws_insights.append([f"Key Finding {i+1}", finding])
            for i, rec in enumerate(insights.get("recommendations", [])):
                ws_insights.append([f"Recommendation {i+1}", rec])

        # Sheet 3: Metadata
        ws_meta = wb.create_sheet(title="Metadata")
        ws_meta.append(["Key", "Value"])
        ws_meta.cell(row=1, column=1).font = Font(bold=True)
        ws_meta.cell(row=1, column=2).font = Font(bold=True)
        ws_meta.column_dimensions['A'].width = 30
        ws_meta.column_dimensions['B'].width = 80
        for k, v in metadata.items():
            ws_meta.append([str(k), str(v)])

        wb.save(file_path)
        return str(file_path)

    @staticmethod
    async def generate_pdf(file_id: uuid.UUID, data: List[Dict[str, Any]], insights: Dict[str, Any], metadata: Dict[str, Any], chart_b64: str = None) -> str:
        file_path = EXPORT_DIR / f"{file_id}.pdf"
        
        class ReportPDF(FPDF):
            def header(self):
                self.set_font("helvetica", "B", 15)
                self.cell(0, 10, metadata.get("workspace_name", "Workspace") + " Report", border=False, ln=1, align="C")
                self.ln(5)
            def footer(self):
                self.set_y(-15)
                self.set_font("helvetica", "I", 8)
                self.cell(0, 10, f"Page {self.page_no()}", align="C")

        pdf = ReportPDF()
        pdf.add_page()
        
        pdf.set_font("helvetica", "B", 12)
        title = metadata.get("saved_query_name") or "Analytics Export"
        pdf.cell(0, 10, title, ln=1)
        pdf.set_font("helvetica", "", 10)
        pdf.cell(0, 8, f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')}", ln=1)
        if metadata.get("user_question"):
            pdf.ln(2)
            pdf.set_font("helvetica", "I", 10)
            pdf.multi_cell(0, 8, f"Question: {metadata['user_question']}")
        pdf.ln(5)

        # SQL
        if metadata.get("generated_sql"):
            pdf.set_font("helvetica", "B", 11)
            pdf.cell(0, 8, "SQL Query", ln=1)
            pdf.set_font("courier", "", 9)
            pdf.set_fill_color(240, 240, 240)
            pdf.multi_cell(0, 6, metadata["generated_sql"], fill=True)
            pdf.ln(5)

        # Chart
        if chart_b64:
            try:
                # Remove data:image/png;base64, prefix if present
                if "," in chart_b64:
                    chart_b64 = chart_b64.split(",")[1]
                img_data = base64.b64decode(chart_b64)
                img_path = EXPORT_DIR / f"{file_id}_chart.png"
                with open(img_path, "wb") as f:
                    f.write(img_data)
                
                pdf.set_font("helvetica", "B", 11)
                pdf.cell(0, 8, "Visualization", ln=1)
                # Keep aspect ratio, max width 180
                pdf.image(str(img_path), w=180)
                pdf.ln(5)
                # Cleanup temp image
                if img_path.exists():
                    img_path.unlink()
            except Exception as e:
                pdf.set_font("helvetica", "I", 9)
                pdf.cell(0, 8, f"Failed to render chart image: {e}", ln=1)

        # Insights
        if insights and insights.get("summary"):
            pdf.set_font("helvetica", "B", 11)
            pdf.cell(0, 8, "AI Insights", ln=1)
            pdf.set_font("helvetica", "", 10)
            pdf.multi_cell(0, 6, insights.get("summary", ""))
            pdf.ln(3)
            for f in insights.get("key_findings", []):
                pdf.multi_cell(0, 6, f"- {f}")
            pdf.ln(5)

        # Table
        if data:
            pdf.set_font("helvetica", "B", 11)
            pdf.cell(0, 8, f"Data Preview (First 50 rows, total: {metadata.get('row_count', len(data))})", ln=1)
            pdf.set_font("helvetica", "", 8)
            headers = list(data[0].keys())[:8] # limit columns to fit
            
            # Print Headers
            pdf.set_font("helvetica", "B", 8)
            col_width = 190 / len(headers)
            for h in headers:
                pdf.cell(col_width, 8, str(h)[:20], border=1)
            pdf.ln()
            
            # Print Data
            pdf.set_font("helvetica", "", 8)
            for row in data[:50]:
                for h in headers:
                    val = str(row.get(h, ""))[:25]
                    pdf.cell(col_width, 8, val, border=1)
                pdf.ln()

        pdf.output(str(file_path))
        return str(file_path)

    @staticmethod
    async def generate_png(file_id: uuid.UUID, chart_b64: str) -> str:
        file_path = EXPORT_DIR / f"{file_id}.png"
        if "," in chart_b64:
            chart_b64 = chart_b64.split(",")[1]
        img_data = base64.b64decode(chart_b64)
        with open(file_path, "wb") as f:
            f.write(img_data)
        return str(file_path)
